#!/usr/bin/env python3
"""Build Changeis Opportunity Intelligence Database Excel workbook."""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.hyperlink import Hyperlink

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
NAVY = "1B2A4A"
WHITE = "FFFFFF"
ALT_BLUE = "C5D9E8"
LIGHT_GRAY = "F2F2F2"
OUTPUT_PATH = "/home/user/workspace/changeis_opportunity_database.xlsx"

HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=18, color=NAVY)
SUBTITLE_FONT = Font(name="Calibri", size=12, color="5A6A7A")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color=NAVY)
LINK_FONT = Font(name="Calibri", size=11, color="0563C1", underline="single")

WHITE_FILL = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
ALT_FILL = PatternFill(start_color=ALT_BLUE, end_color=ALT_BLUE, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

# Priority NAICS codes for IT/consulting
PRIORITY_NAICS = {
    "541511", "541512", "541513", "541519",
    "541611", "541612", "541613", "541614", "541618",
    "541715", "541990",
}

# Full list of 46 Changeis-relevant NAICS codes with descriptions
CHANGEIS_NAICS = {
    "541511": "Custom Computer Programming Services",
    "541512": "Computer Systems Design Services",
    "541513": "Computer Facilities Management Services",
    "541519": "Other Computer Related Services",
    "541611": "Administrative Management and General Management Consulting Services",
    "541612": "Human Resources Consulting Services",
    "541613": "Marketing Consulting Services",
    "541614": "Process, Physical Distribution, and Logistics Consulting Services",
    "541618": "Other Management Consulting Services",
    "541620": "Environmental Consulting Services",
    "541690": "Other Scientific and Technical Consulting Services",
    "541715": "Research and Development in the Physical, Engineering, and Life Sciences (except Nanotechnology and Biotechnology)",
    "541720": "Research and Development in the Social Sciences and Humanities",
    "541990": "All Other Professional, Scientific, and Technical Services",
    "541330": "Engineering Services",
    "541380": "Testing Laboratories",
    "541410": "Interior Design Services",
    "541420": "Industrial Design Services",
    "541430": "Graphic Design Services",
    "541490": "Other Specialized Design Services",
    "541810": "Advertising Agencies",
    "541820": "Public Relations Agencies",
    "541830": "Media Buying Agencies",
    "541840": "Media Representatives",
    "541850": "Outdoor Advertising",
    "541860": "Direct Mail Advertising",
    "541870": "Advertising Material Distribution Services",
    "541890": "Other Services Related to Advertising",
    "541910": "Marketing Research and Public Opinion Polling",
    "541921": "Photography Studios, Portrait",
    "541922": "Commercial Photography",
    "518210": "Data Processing, Hosting, and Related Services",
    "519130": "Internet Publishing and Broadcasting and Web Search Portals",
    "519190": "All Other Information Services",
    "561110": "Office Administrative Services",
    "561210": "Facilities Support Services",
    "561310": "Employment Placement Agencies",
    "561320": "Temporary Help Services",
    "561330": "Professional Employer Organizations",
    "561410": "Document Preparation Services",
    "561422": "Telemarketing Bureaus and Other Contact Centers",
    "561439": "Other Business Service Centers (including Copy Shops)",
    "561499": "All Other Business Support Services",
    "561611": "Investigation Services",
    "561612": "Security Guards and Patrol Services",
    "611420": "Computer Training",
    "5417": "Scientific Research and Development Services",
}

# Dollar range sort order (descending by value)
DOLLAR_RANGE_ORDER = {
    "Over $100M": 0,
    "$50M to $100M": 1,
    "$25M to $50M": 2,
    "$10M to $25M": 3,
    "$5M to $10M": 4,
    "$2M to $5M": 5,
    "$1M to $2M": 6,
    "$500K to $1M": 7,
    "$350K to $500K": 8,
    "$150K to $350K": 9,
    "Under $150K": 10,
    "TBD": 11,
    "N/A": 12,
    "": 13,
}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_data():
    with open("/home/user/workspace/apfs_changeis_filtered.json") as f:
        apfs_data = json.load(f)
    with open("/home/user/workspace/sam_gov_data.json") as f:
        sam_data = json.load(f)
    return apfs_data, sam_data


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def set_col_widths(ws, widths):
    """Set column widths from a dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def apply_body_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        fill = WHITE_FILL if (r - start_row) % 2 == 0 else ALT_FILL
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_excel_table(ws, ref, name, show_filters=True):
    """Add an Excel Table object to the worksheet."""
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = style
    if not show_filters:
        table.autoFilter = None
    ws.add_table(table)


def extract_naics_code(naics_str):
    """Extract just the NAICS code from a string like '541511 - Custom Computer Programming'."""
    if not naics_str:
        return ""
    return naics_str.split(" ")[0].strip() if naics_str else ""


def is_priority_naics(naics_str):
    code = extract_naics_code(naics_str)
    return code in PRIORITY_NAICS


def get_dollar_sort_key(dollar_range_str):
    return DOLLAR_RANGE_ORDER.get(dollar_range_str, 12)


def extract_relevance_keyword(text):
    """Extract relevance level from text like 'HIGHEST — ...' or 'HIGH — ...'."""
    if not text:
        return ""
    text_upper = text.upper()
    for level in ["HIGHEST", "HIGH", "MEDIUM-HIGH", "MEDIUM", "LOW-MEDIUM", "LOW"]:
        if text_upper.startswith(level):
            return level
    return ""


# ---------------------------------------------------------------------------
# Build unified rows for "All Opportunities" sheet
# ---------------------------------------------------------------------------
def build_unified_rows(apfs_data, sam_data):
    rows = []

    # APFS records
    for rec in apfs_data:
        dollar_display = rec.get("dollar_range", {})
        if isinstance(dollar_display, dict):
            dollar_str = dollar_display.get("display_name", "")
        else:
            dollar_str = str(dollar_display) if dollar_display else ""

        naics_str = rec.get("naics", "")
        naics_code = extract_naics_code(naics_str)

        rows.append({
            "source": "DHS APFS",
            "agency": "DHS",
            "sub_agency": rec.get("organization", ""),
            "title": rec.get("requirements_title", ""),
            "naics_code": naics_code,
            "dollar_range": dollar_str,
            "contract_type": rec.get("contract_type", ""),
            "set_aside": rec.get("small_business_program", "None") or "None",
            "contract_vehicle": rec.get("contract_vehicle", ""),
            "status": "Forecast",
            "award_quarter": rec.get("award_quarter", ""),
            "release_date": rec.get("estimated_release_date", ""),
            "published_date": rec.get("publish_date", ""),
            "relevance": "Priority" if is_priority_naics(naics_str) else "",
            "url": "",
            "_sort_naics": 0 if is_priority_naics(naics_str) else 1,
            "_sort_dollar": get_dollar_sort_key(dollar_str),
        })

    # SAM.gov high-relevance opportunities
    for opp in sam_data.get("high_relevance_opportunities", []):
        naics_str = opp.get("naics_code", "")
        naics_code = extract_naics_code(naics_str)
        relevance = extract_relevance_keyword(opp.get("relevance_to_changeis", ""))

        rows.append({
            "source": "SAM.gov",
            "agency": "DOT",
            "sub_agency": "FAA",
            "title": opp.get("title", ""),
            "naics_code": naics_code,
            "dollar_range": "",
            "contract_type": "",
            "set_aside": opp.get("set_aside", ""),
            "contract_vehicle": "",
            "status": "Active",
            "award_quarter": "",
            "release_date": opp.get("response_deadline", ""),
            "published_date": "",
            "relevance": relevance,
            "url": opp.get("sam_url", ""),
            "_sort_naics": 0 if is_priority_naics(naics_str) else 1,
            "_sort_dollar": 12,
        })

    # SAM.gov lower-relevance
    for opp in sam_data.get("lower_relevance_opportunities", []):
        rows.append({
            "source": "SAM.gov",
            "agency": "DOT",
            "sub_agency": "FAA",
            "title": opp.get("title", ""),
            "naics_code": "",
            "dollar_range": "",
            "contract_type": "",
            "set_aside": "",
            "contract_vehicle": "",
            "status": "Active",
            "award_quarter": "",
            "release_date": opp.get("response_deadline", ""),
            "published_date": "",
            "relevance": "LOW",
            "url": "",
            "_sort_naics": 2,
            "_sort_dollar": 12,
        })

    return rows


# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------
def build_overview(wb, apfs_data, sam_data, unified_rows):
    ws = wb.active
    ws.title = "Overview"
    ws.sheet_properties.tabColor = NAVY

    # Margins per SKILL.md: Column A empty width 3, Row 1 empty
    ws.column_dimensions["A"].width = 3

    # Title
    ws.merge_cells("B2:H2")
    cell = ws["B2"]
    cell.value = "Changeis Opportunity Intelligence Database"
    cell.font = TITLE_FONT
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[2].height = 35

    # Subtitle
    ws.merge_cells("B3:H3")
    cell = ws["B3"]
    cell.value = "Consolidated FAA SAM.gov + DHS APFS Opportunities"
    cell.font = SUBTITLE_FONT
    ws.row_dimensions[3].height = 22

    # Generated date
    ws.merge_cells("B4:H4")
    cell = ws["B4"]
    cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    cell.font = Font(name="Calibri", size=10, italic=True, color="888888")
    ws.row_dimensions[4].height = 18

    # Summary stats
    row = 6
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = "Summary Statistics"
    ws[f"B{row}"].font = SECTION_FONT
    ws.row_dimensions[row].height = 25

    total_apfs = len(apfs_data)
    total_sam_high = len(sam_data.get("high_relevance_opportunities", []))
    total_sam_low = len(sam_data.get("lower_relevance_opportunities", []))
    total_sam = total_sam_high + total_sam_low
    total_all = total_apfs + total_sam

    # Count by dollar range
    dollar_counts = {}
    for rec in apfs_data:
        dr = rec.get("dollar_range", {})
        name = dr.get("display_name", "N/A") if isinstance(dr, dict) else "N/A"
        dollar_counts[name] = dollar_counts.get(name, 0) + 1

    # Count by organization
    org_counts = {}
    for rec in apfs_data:
        org = rec.get("organization", "Unknown")
        org_counts[org] = org_counts.get(org, 0) + 1

    stats = [
        ("Total Opportunities", total_all),
        ("DHS APFS Opportunities", total_apfs),
        ("FAA SAM.gov Opportunities (High Relevance)", total_sam_high),
        ("FAA SAM.gov Opportunities (Lower Relevance)", total_sam_low),
    ]

    row = 8
    label_font = Font(name="Calibri", size=11)
    value_font = Font(name="Calibri", size=11, bold=True, color=NAVY)
    for label, value in stats:
        ws[f"B{row}"].value = label
        ws[f"B{row}"].font = label_font
        ws[f"D{row}"].value = value
        ws[f"D{row}"].font = value_font
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        row += 1

    # Dollar range breakdown
    row += 1
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = "DHS APFS by Dollar Range"
    ws[f"B{row}"].font = SECTION_FONT
    ws.row_dimensions[row].height = 25
    row += 1

    sorted_dollars = sorted(dollar_counts.items(), key=lambda x: DOLLAR_RANGE_ORDER.get(x[0], 99))
    for name, count in sorted_dollars:
        ws[f"B{row}"].value = name
        ws[f"B{row}"].font = label_font
        ws[f"D{row}"].value = count
        ws[f"D{row}"].font = value_font
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        row += 1

    # Agency/org breakdown
    row += 1
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = "DHS APFS by Organization"
    ws[f"B{row}"].font = SECTION_FONT
    ws.row_dimensions[row].height = 25
    row += 1

    sorted_orgs = sorted(org_counts.items(), key=lambda x: -x[1])
    for org, count in sorted_orgs:
        ws[f"B{row}"].value = org
        ws[f"B{row}"].font = label_font
        ws[f"D{row}"].value = count
        ws[f"D{row}"].font = value_font
        ws[f"D{row}"].alignment = Alignment(horizontal="right")
        row += 1

    # Navigation links
    row += 2
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = "Quick Navigation"
    ws[f"B{row}"].font = SECTION_FONT
    ws.row_dimensions[row].height = 25
    row += 1

    sheets_info = [
        ("All Opportunities", "Unified view of all opportunities"),
        ("DHS APFS Detail", "Full DHS APFS dataset with all fields"),
        ("FAA SAM.gov Detail", "FAA SAM.gov opportunity details"),
        ("NAICS Reference", "Changeis-relevant NAICS codes"),
    ]
    for sheet_name, description in sheets_info:
        cell = ws[f"B{row}"]
        cell.value = sheet_name
        cell.font = LINK_FONT
        cell.hyperlink = f"#'{sheet_name}'!A1"
        ws[f"C{row}"].value = description
        ws[f"C{row}"].font = Font(name="Calibri", size=10, color="666666")
        row += 1

    set_col_widths(ws, {"B": 45, "C": 45, "D": 15, "E": 15})


def build_all_opportunities(wb, unified_rows):
    ws = wb.create_sheet("All Opportunities")
    ws.sheet_properties.tabColor = NAVY

    headers = [
        "Source", "Agency", "Sub-Agency/Organization", "Title",
        "NAICS Code", "Dollar Range", "Contract Type",
        "Set-Aside/Small Business", "Contract Vehicle", "Status",
        "Award Quarter / Release Date", "Published Date",
        "Relevance", "URL/Link"
    ]

    col_widths = {
        "A": 12, "B": 10, "C": 22, "D": 50,
        "E": 14, "F": 18, "G": 18,
        "H": 25, "I": 25, "J": 12,
        "K": 28, "L": 15, "M": 14, "N": 45,
    }
    set_col_widths(ws, col_widths)

    # Write headers
    header_row = 1
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c, value=h)
    apply_header_style(ws, header_row, len(headers))
    ws.row_dimensions[header_row].height = 25

    # Sort: priority NAICS first, then dollar range descending
    sorted_rows = sorted(unified_rows, key=lambda r: (r["_sort_naics"], r["_sort_dollar"]))

    # Write data
    for i, row_data in enumerate(sorted_rows):
        r = i + 2
        ws.cell(row=r, column=1, value=row_data["source"])
        ws.cell(row=r, column=2, value=row_data["agency"])
        ws.cell(row=r, column=3, value=row_data["sub_agency"])
        ws.cell(row=r, column=4, value=row_data["title"])
        ws.cell(row=r, column=5, value=row_data["naics_code"])
        ws.cell(row=r, column=6, value=row_data["dollar_range"])
        ws.cell(row=r, column=7, value=row_data["contract_type"])
        ws.cell(row=r, column=8, value=row_data["set_aside"])
        ws.cell(row=r, column=9, value=row_data["contract_vehicle"])
        ws.cell(row=r, column=10, value=row_data["status"])

        # Combine award quarter and release date
        aq = row_data.get("award_quarter", "")
        rd = row_data.get("release_date", "")
        combined = f"{aq} / {rd}".strip(" /") if aq or rd else ""
        ws.cell(row=r, column=11, value=combined)

        ws.cell(row=r, column=12, value=row_data["published_date"])
        ws.cell(row=r, column=13, value=row_data["relevance"])

        url = row_data.get("url", "")
        if url:
            cell = ws.cell(row=r, column=14, value=url)
            cell.hyperlink = url
            cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        else:
            ws.cell(row=r, column=14, value="")

    last_row = len(sorted_rows) + 1
    apply_body_rows(ws, 2, last_row, len(headers))

    # Excel Table
    ref = f"A1:{get_column_letter(len(headers))}{last_row}"
    add_excel_table(ws, ref, "AllOpportunities")

    # Freeze panes
    ws.freeze_panes = "A2"


def build_apfs_detail(wb, apfs_data):
    ws = wb.create_sheet("DHS APFS Detail")
    ws.sheet_properties.tabColor = "2E5090"

    headers = [
        "APFS Number", "Organization", "Requirements Title", "NAICS",
        "Dollar Range", "Contract Type", "Contract Vehicle",
        "Small Business Program", "Competitive", "Award Quarter",
        "Est. Release Date", "Publish Date", "Contract Status",
        "Requirement Description", "Requirements Office", "Contracting Office",
        "Est. PoP Start", "Est. PoP End", "Anticipated Award Date",
        "Place of Performance", "Requirements Contact",
        "Requirements Contact Email", "Requirements Contact Phone",
        "Alternate Contact", "Alternate Contact Email",
        "SBS Coordinator", "SBS Coordinator Email",
        "Fiscal Year", "Created On", "Last Updated",
    ]

    # Write headers
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 25

    # Sort: priority NAICS first, then by dollar range descending
    def apfs_sort_key(rec):
        naics = rec.get("naics", "")
        is_priority = 0 if is_priority_naics(naics) else 1
        dr = rec.get("dollar_range", {})
        dollar_name = dr.get("display_name", "") if isinstance(dr, dict) else ""
        return (is_priority, get_dollar_sort_key(dollar_name))

    sorted_apfs = sorted(apfs_data, key=apfs_sort_key)

    for i, rec in enumerate(sorted_apfs):
        r = i + 2
        dr = rec.get("dollar_range", {})
        dollar_name = dr.get("display_name", "") if isinstance(dr, dict) else ""

        pop_city = rec.get("place_of_performance_city", "")
        pop_state = rec.get("place_of_performance_state", "")
        pop = f"{pop_city}, {pop_state}".strip(", ") if pop_city or pop_state else ""

        req_contact = f"{rec.get('requirements_contact_first_name', '')} {rec.get('requirements_contact_last_name', '')}".strip()
        alt_contact = f"{rec.get('alternate_contact_first_name', '')} {rec.get('alternate_contact_last_name', '')}".strip()
        sbs_contact = f"{rec.get('sbs_coordinator_first_name', '')} {rec.get('sbs_coordinator_last_name', '')}".strip()

        values = [
            rec.get("apfs_number", ""),
            rec.get("organization", ""),
            rec.get("requirements_title", ""),
            rec.get("naics", ""),
            dollar_name,
            rec.get("contract_type", ""),
            rec.get("contract_vehicle", ""),
            rec.get("small_business_program", "") or "None",
            rec.get("competitive", ""),
            rec.get("award_quarter", ""),
            rec.get("estimated_release_date", ""),
            rec.get("publish_date", ""),
            rec.get("contract_status", ""),
            rec.get("requirement", ""),
            rec.get("requirements_office", ""),
            rec.get("contracting_office", ""),
            rec.get("estimated_period_of_performance_start", ""),
            rec.get("estimated_period_of_performance_end", ""),
            rec.get("anticipated_award_date", ""),
            pop,
            req_contact,
            rec.get("requirements_contact_email", ""),
            rec.get("requirements_contact_phone", ""),
            alt_contact,
            rec.get("alternate_contact_email", ""),
            sbs_contact,
            rec.get("sbs_coordinator_email", ""),
            rec.get("fiscal_year", ""),
            rec.get("created_on", ""),
            rec.get("last_updated_date", ""),
        ]

        for c, val in enumerate(values, 1):
            ws.cell(row=r, column=c, value=val)

    last_row = len(sorted_apfs) + 1
    apply_body_rows(ws, 2, last_row, len(headers))

    # Column widths
    widths = {
        "A": 18, "B": 12, "C": 45, "D": 38,
        "E": 18, "F": 18, "G": 25,
        "H": 22, "I": 30, "J": 14,
        "K": 16, "L": 14, "M": 14,
        "N": 60, "O": 35, "P": 35,
        "Q": 16, "R": 16, "S": 18,
        "T": 22, "U": 22, "V": 32,
        "W": 18, "X": 22, "Y": 32,
        "Z": 22, "AA": 32, "AB": 12,
        "AC": 14, "AD": 14,
    }
    set_col_widths(ws, widths)

    # Excel Table
    last_col_letter = get_column_letter(len(headers))
    ref = f"A1:{last_col_letter}{last_row}"
    add_excel_table(ws, ref, "DHSAPFSDetail")

    ws.freeze_panes = "A2"


def build_sam_detail(wb, sam_data):
    ws = wb.create_sheet("FAA SAM.gov Detail")
    ws.sheet_properties.tabColor = "1A5276"

    headers = [
        "Title", "Solicitation Number", "Notice Type", "NAICS Code",
        "PSC Code", "Set-Aside", "Response Deadline",
        "Place of Performance", "Contact", "Office",
        "Description", "Relevance to Changeis", "Status Note",
        "SAM.gov URL", "Relevance Level",
    ]

    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    apply_header_style(ws, 1, len(headers))
    ws.row_dimensions[1].height = 25

    row_num = 2

    # High relevance first
    for opp in sam_data.get("high_relevance_opportunities", []):
        relevance_level = extract_relevance_keyword(opp.get("relevance_to_changeis", ""))
        values = [
            opp.get("title", ""),
            opp.get("solicitation_number", ""),
            opp.get("notice_type", ""),
            opp.get("naics_code", ""),
            opp.get("psc", ""),
            opp.get("set_aside", ""),
            opp.get("response_deadline", ""),
            opp.get("place_of_performance", ""),
            opp.get("contact", ""),
            opp.get("office", ""),
            opp.get("description", ""),
            opp.get("relevance_to_changeis", ""),
            opp.get("status_note", ""),
            opp.get("sam_url", ""),
            relevance_level,
        ]
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=c, value=val)
            if c == 14 and val:  # URL column
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        row_num += 1

    # Lower relevance
    for opp in sam_data.get("lower_relevance_opportunities", []):
        values = [
            opp.get("title", ""),
            opp.get("solicitation_number", ""),
            opp.get("notice_type", ""),
            "", "",
            "",
            opp.get("response_deadline", ""),
            "", "", "",
            "",
            opp.get("relevance", ""),
            "",
            "",
            "LOW",
        ]
        for c, val in enumerate(values, 1):
            ws.cell(row=row_num, column=c, value=val)
        row_num += 1

    last_row = row_num - 1
    apply_body_rows(ws, 2, last_row, len(headers))

    widths = {
        "A": 50, "B": 30, "C": 28, "D": 38,
        "E": 40, "F": 30, "G": 32,
        "H": 25, "I": 35, "J": 30,
        "K": 60, "L": 50, "M": 35,
        "N": 50, "O": 16,
    }
    set_col_widths(ws, widths)

    last_col_letter = get_column_letter(len(headers))
    ref = f"A1:{last_col_letter}{last_row}"
    add_excel_table(ws, ref, "FAASAMDetail")

    ws.freeze_panes = "A2"


def build_naics_reference(wb):
    ws = wb.create_sheet("NAICS Reference")
    ws.sheet_properties.tabColor = "1B4F72"

    # Column A margin
    ws.column_dimensions["A"].width = 3

    # Title
    ws.merge_cells("B1:D1")
    ws["B1"].value = "Changeis-Relevant NAICS Codes"
    ws["B1"].font = Font(name="Calibri", bold=True, size=14, color=NAVY)
    ws.row_dimensions[1].height = 30

    ws["B2"].value = f"{len(CHANGEIS_NAICS)} NAICS codes relevant to Changeis capabilities"
    ws["B2"].font = Font(name="Calibri", size=10, italic=True, color="888888")

    # Headers at row 4
    header_row = 4
    headers = ["NAICS Code", "Description", "Priority IT/Consulting"]
    header_cols = {"B": headers[0], "C": headers[1], "D": headers[2]}
    for col_letter, header_text in header_cols.items():
        cell = ws[f"{col_letter}{header_row}"]
        cell.value = header_text
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 25

    # Data
    sorted_naics = sorted(
        CHANGEIS_NAICS.items(),
        key=lambda x: (0 if x[0] in PRIORITY_NAICS else 1, x[0])
    )

    highlight_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

    for i, (code, desc) in enumerate(sorted_naics):
        r = header_row + 1 + i
        is_priority = code in PRIORITY_NAICS
        fill = highlight_fill if is_priority else (WHITE_FILL if i % 2 == 0 else ALT_FILL)

        ws[f"B{r}"].value = code
        ws[f"B{r}"].font = Font(name="Calibri", size=10, bold=is_priority)
        ws[f"B{r}"].fill = fill
        ws[f"B{r}"].border = THIN_BORDER

        ws[f"C{r}"].value = desc
        ws[f"C{r}"].font = Font(name="Calibri", size=10, bold=is_priority)
        ws[f"C{r}"].fill = fill
        ws[f"C{r}"].border = THIN_BORDER

        ws[f"D{r}"].value = "Yes" if is_priority else ""
        ws[f"D{r}"].font = Font(
            name="Calibri", size=10, bold=True,
            color="2E7D32" if is_priority else "000000"
        )
        ws[f"D{r}"].fill = fill
        ws[f"D{r}"].border = THIN_BORDER
        ws[f"D{r}"].alignment = Alignment(horizontal="center")

    last_row = header_row + len(CHANGEIS_NAICS)
    ref = f"B{header_row}:D{last_row}"
    add_excel_table(ws, ref, "NAICSReference")

    set_col_widths(ws, {"B": 16, "C": 60, "D": 22})
    ws.freeze_panes = f"A{header_row + 1}"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading data...")
    apfs_data, sam_data = load_data()

    print(f"  APFS records: {len(apfs_data)}")
    print(f"  SAM.gov high relevance: {len(sam_data.get('high_relevance_opportunities', []))}")
    print(f"  SAM.gov lower relevance: {len(sam_data.get('lower_relevance_opportunities', []))}")

    unified_rows = build_unified_rows(apfs_data, sam_data)
    print(f"  Unified rows: {len(unified_rows)}")

    wb = Workbook()

    print("Building Overview sheet...")
    build_overview(wb, apfs_data, sam_data, unified_rows)

    print("Building All Opportunities sheet...")
    build_all_opportunities(wb, unified_rows)

    print("Building DHS APFS Detail sheet...")
    build_apfs_detail(wb, apfs_data)

    print("Building FAA SAM.gov Detail sheet...")
    build_sam_detail(wb, sam_data)

    print("Building NAICS Reference sheet...")
    build_naics_reference(wb)

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print("Done!")


if __name__ == "__main__":
    main()
